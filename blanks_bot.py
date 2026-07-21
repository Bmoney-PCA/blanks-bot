#!/usr/bin/env python3
# =============================================================================
#  blanks_bot.py  —  Printavo → SanMar blanks ordering automation
#  SINGLE-FILE HANDOFF.  Read this whole header before changing anything.
# =============================================================================
#
#  WHO THIS IS FOR
#  ---------------
#  You're picking up a half-built automation for a screen-printing shop
#  (Pure Choice Apparel). You can use Claude (or any assistant) to finish the
#  parts marked `TODO(next-builder)`. Everything outside those TODOs is built
#  and unit-tested against the real Printavo/SanMar API docs. Don't rewrite it.
#
#  WHAT THE SYSTEM DOES (the business goal)
#  ----------------------------------------
#  Orders flow:  Shopify → Printflo → Printavo. When an order lands in a
#  specific Printavo status (the "needs blanks" bucket), the shop must order
#  blank garments from SanMar (wholesaler) by item# + color + size. This bot:
#
#    1. Pulls EVERY order currently in the configured Printavo status
#       (status-driven, not date-driven — a rolling accumulation).
#    2. Sums identical (style#, color, size) across all those orders so blanks
#       are ordered ONCE in bulk.
#    3. Maps Printavo (style, color) -> SanMar (style, color code) via a CSV the
#       shop controls. Anything unmapped/unconfirmed is FLAGGED, never ordered.
#    4. Builds an XLSX "Blanks To Order" sheet (2 tabs: order / needs-attention)
#       and emails it -- designed to run on a 3 PM cron.
#    5. On explicit approval, submits the SanMar purchase order, then writes
#       back to Printavo: moves each fully-ordered order to "AWAITING GARMENTS"
#       and checks its "Order Blanks" task. This also prevents double-ordering,
#       because advanced orders leave the pull bucket.
#
#  ARCHITECTURE
#  ------------
#    Printavo (GraphQL, pull by statusIds)
#          |  orders -> line items -> (style#, color, size, qty)
#          v
#    aggregate_demand()  -> sums identical SKUs across all orders in the status
#          |
#          v
#    sku_map.csv  -> Printavo (style,color) -> SanMar (style, colorCode)
#          |  unmapped/unconfirmed -> "Needs Attention" tab, never ordered
#          v
#    build_report() -> XLSX  ->  email_report() (3 PM cron)
#          |
#       (human approves: --submit --i-approve, and sanmar.po_enabled: true)
#          v
#    submit_sanmar_po() -> PromoStandards PO (SOAP)
#          v
#    advance_orders() -> statusUpdate -> AWAITING GARMENTS + taskUpdate (Order Blanks done)
#                       SAFETY: only advances an order if EVERY blank line on it
#                       was actually ordered; mixed/held orders stay in place.
#
#  VERIFIED API FACTS (confirmed against live docs, June 2026)
#  -----------------------------------------------------------
#  PRINTAVO API v2 (GraphQL)
#    - Endpoint:  https://www.printavo.com/api/v2
#    - Auth headers:  email: <account email>,  token: <My Account > API token>
#    - Rate limit:  10 requests / 5 seconds. Connections paginate 25 by default.
#    - Query:  orders(statusIds: [ID!], first, after, sortOn) -> union Invoice|Quote
#    - Query:  statuses { nodes { id name } }
#    - Mutation:  statusUpdate(parentId: ID!, statusId: ID!)     # set order status
#    - Mutation:  taskUpdate(id: ID!, input: TaskInput)          # TaskInput.completed: Boolean
#    - Order/Invoice/Quote expose: tasks { nodes { id name completed } }
#
#  SANMAR (PromoStandards PO Service, SOAP)
#    - PROD WSDL:  https://ws.sanmar.com:8080/promostandards/POServiceBinding?WSDL
#    - TEST WSDL:  https://test-ws.sanmar.com:8080/promostandards/POServiceBinding?WSDL
#    - Auth:  your sanmar.com username/password authenticate PromoStandards APIs.
#    - PO integration is GATED: request it separately AFTER product-data access.
#      SanMar provisions a Test env in ~2-3 business days and requires a
#      multi-line test order before enabling production. Start this early --
#      it's the long pole.
#    - Recommended: call getPreSubmitInfo (inventory) BEFORE sendPO. If stock
#      can't fill a line at one warehouse, the whole PO can go on hold for
#      manual key-in. SanMar carries Next Level NL1810 under that style number.
#
#  CURRENT STATE
#  -------------
#  DONE + unit-tested (offline, against the shop's real screenshot data):
#    - Status-based pull + pagination                  fetch_printavo_orders()
#    - Status discovery helper                          list_statuses()
#    - Demand aggregation by (style,color,size)         aggregate_demand()       TESTED
#    - Per-order records + "Order Blanks" task lookup   extract_order_records()
#    - SKU mapping + unmapped/unconfirmed flagging       map_to_sanmar()          TESTED
#    - XLSX report (2 tabs)                              build_report()           TESTED
#    - Email w/ attachment (SMTP)                        email_report()
#    - Printavo write-back w/ coverage safety            advance_orders()         TESTED
#    - SanMar PO payload assembly + DRY RUN              submit_sanmar_po()       TESTED (dry)
#
#  TODO(next-builder)  -- grep for this tag in the code:
#    [A] submit_sanmar_po() LIVE path: the `payload` dict is a placeholder. Build
#        the real request from the WSDL via zeep type factories
#        (client.get_type('ns0:...')) and confirm the operation name (sendPO vs
#        submitPO) + argument shape against the live POServiceBinding WSDL.
#        Test on the TEST WSDL first. Do NOT ship live until a multi-line test
#        PO invoices correctly in SanMar's test env.
#    [B] sanmar_check_stock() getPreSubmitInfo: wire against the live WSDL and
#        add an "Available" column to the report. Currently raises NotImplemented.
#    [C] Verify Printavo v2 size shape on the real account via the API explorer
#        (Account > API). Query assumes lineItems.sizes{ size quantity }. The
#        parser (_normalize_sizes) already tolerates both v2 and legacy shapes,
#        but confirm and delete the branch you don't need.
#    [D] Fill sku_map.csv with the shop's real, CONFIRMED mappings. Note
#        "Gray (HeavyWeight)" likely maps to Next Level's Heavyweight tee (a
#        DIFFERENT SanMar style), not the 1810 -- verify on sanmar.com.
#
#  GOTCHAS
#  -------
#    - AWAITING GARMENTS status must NOT be in printavo.status_ids (the pull
#      list), or advanced orders get re-pulled forever.
#    - Write-backs ONLY fire on a live PO (--submit --i-approve AND
#      sanmar.po_enabled: true). Otherwise advance_orders() just previews.
#    - order_blanks_task_name must match the Printavo preset task name exactly
#      (case-insensitive). If renamed, the task-check silently no-ops.
#    - Money is real. Keep running in DRY RUN for days and eyeball the sheet +
#      the "would advance / held" preview before enabling live submit.
#
#  SETUP
#  -----
#    pip install requests pyyaml openpyxl zeep
#    python blanks_bot.py --init          # writes config.example.yaml + sku_map.csv
#    cp config.example.yaml config.yaml   # then fill in credentials/IDs
#    python blanks_bot.py --list-statuses # find pull status + AWAITING GARMENTS IDs
#
#  RUN MODES
#    python blanks_bot.py --once                      # pull -> sheet -> email (no buying)
#    python blanks_bot.py --once --date 2026-06-17    # snapshot label override
#    python blanks_bot.py --once --check-stock        # (after TODO[B]) add stock column
#    python blanks_bot.py --once --submit             # dry PO + dry write-back preview
#    python blanks_bot.py --once --submit --i-approve # LIVE (also needs po_enabled: true)
#
#  CRON (3 PM daily):
#    0 15 * * *  /usr/bin/python3 /path/blanks_bot.py --once >> /var/log/blanks_bot.log 2>&1
# =============================================================================

import argparse
import csv
import datetime as dt
import smtplib
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from email.message import EmailMessage
from pathlib import Path

import requests
import yaml

# ----------------------------------------------------------------------------
# Embedded support files (written to disk by `--init`)
# ----------------------------------------------------------------------------

CONFIG_EXAMPLE = """\
printavo:
  email: you@purechoiceapparel.com      # the email tied to your Printavo API token
  token: PRINTAVO_API_TOKEN             # My Account > API
  status_ids:                           # the status Printflo/Shopify pushes into.
    - "0000"                            # run `--list-statuses` to find your real ID(s)
  awaiting_garments_status_id: "0000"   # status to move orders to AFTER blanks ordered
  order_blanks_task_name: "Order Blanks"  # task name to auto-check on PO submit

sku_map_path: sku_map.csv
output_dir: ./reports

email:
  enabled: true
  smtp_host: smtp.gmail.com
  smtp_port: 587
  smtp_user: alerts@yourdomain.com
  smtp_pass: APP_PASSWORD               # use an app password, not your login
  from: alerts@yourdomain.com
  to:
    - you@purechoiceapparel.com

sanmar:
  po_enabled: false                     # flip to true ONLY after SanMar PO onboarding
  username: SANMAR_WEB_USER             # your sanmar.com username (not Printavo)
  password: SANMAR_WEB_PASS
  customer_number: "0000000"
  po_wsdl: "https://ws.sanmar.com:8080/promostandards/POServiceBinding?WSDL"
  ship_to_name: "Pure Choice Apparel"
  ship_to_address: "123 Shop St, Phoenix, AZ 85001"
"""

SKU_MAP_EXAMPLE = """\
printavo_style,printavo_color,sanmar_style,sanmar_color,confirmed
NL1810,Black,NL1810,Black,yes
NL1810,Gray (HeavyWeight),NL1810,Heather Gray,no
"""


def write_init_files():
    """Write config.example.yaml + sku_map.csv if absent. Never overwrites."""
    for name, content in [("config.example.yaml", CONFIG_EXAMPLE),
                          ("sku_map.csv", SKU_MAP_EXAMPLE)]:
        if Path(name).exists():
            print(f"[init] {name} already exists -- left untouched")
        else:
            Path(name).write_text(content)
            print(f"[init] wrote {name}")
    print("[init] next: cp config.example.yaml config.yaml, fill it in, "
          "then `--list-statuses`.")


# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------

SIZE_ORDER = ["XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL", "6XL"]

# Maps Printavo's internal size field names (legacy + v2 variants) to our labels.
SIZE_ALIASES = {
    "xs": "XS", "size_xs": "XS",
    "s": "S", "size_s": "S",
    "m": "M", "size_m": "M",
    "l": "L", "size_l": "L",
    "xl": "XL", "size_xl": "XL",
    "2xl": "2XL", "xxl": "2XL", "size_2xl": "2XL",
    "3xl": "3XL", "xxxl": "3XL", "size_3xl": "3XL",
    "4xl": "4XL", "size_4xl": "4XL",
    "5xl": "5XL", "size_5xl": "5XL",
    "6xl": "6XL", "size_6xl": "6XL",
}

PRINTAVO_ENDPOINT = "https://www.printavo.com/api/v2"


def load_config(path="config.yaml"):
    if not Path(path).exists():
        sys.exit(f"Missing {path}. Run `python blanks_bot.py --init`, then "
                 f"copy config.example.yaml to config.yaml and fill it in.")
    with open(path) as f:
        return yaml.safe_load(f)


# ----------------------------------------------------------------------------
# 1. Printavo pull
# ----------------------------------------------------------------------------

PRINTAVO_STATUSES_QUERY = "query { statuses { nodes { id name } } }"

# TODO(next-builder)[C]: verify lineItems.sizes shape on the real account.
PRINTAVO_ORDERS_QUERY = """
query Orders($after: String, $statusIds: [ID!]) {
  orders(first: 25, after: $after, statusIds: $statusIds,
         sortOn: CREATED_AT_DESC) {
    pageInfo { hasNextPage endCursor }
    nodes {
      ... on Invoice {
        id visualId
        tasks { nodes { id name completed } }
        lineItemGroups {
          nodes {
            lineItems {
              nodes { id styleNumber styleDescription color sizes { size quantity } }
            }
          }
        }
      }
      ... on Quote {
        id visualId
        tasks { nodes { id name completed } }
        lineItemGroups {
          nodes {
            lineItems {
              nodes { id styleNumber styleDescription color sizes { size quantity } }
            }
          }
        }
      }
    }
  }
}
"""


def printavo_headers(cfg):
    return {
        "Content-Type": "application/json",
        "email": cfg["printavo"]["email"],
        "token": cfg["printavo"]["token"],
    }


def list_statuses(cfg):
    """Print all Printavo statuses + IDs so you can find your 'blanks needed'
    status AND your AWAITING GARMENTS status."""
    resp = requests.post(PRINTAVO_ENDPOINT, headers=printavo_headers(cfg),
                         json={"query": PRINTAVO_STATUSES_QUERY}, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if "errors" in data:
        raise RuntimeError(f"Printavo GraphQL error: {data['errors']}")
    statuses = data["data"]["statuses"]
    nodes = statuses.get("nodes", statuses) if isinstance(statuses, dict) else statuses
    print("Printavo statuses (status_ids = pull bucket; "
          "awaiting_garments_status_id = post-order):")
    for s in nodes:
        print(f"  {s['id']:>8}  {s['name']}")
    return nodes


def fetch_printavo_orders(cfg, status_ids):
    """Return ALL order nodes currently in the given status(es). Paginated.

    Status-driven, not date-driven: every order Printflo/Shopify has pushed into
    these statuses is accumulated, regardless of when it arrived.
    """
    if not status_ids:
        raise SystemExit("No status_ids configured. Run --list-statuses first, "
                         "then set printavo.status_ids in config.yaml.")
    status_ids = [str(s) for s in status_ids]
    after, nodes = None, []
    while True:
        resp = requests.post(
            PRINTAVO_ENDPOINT,
            headers=printavo_headers(cfg),
            json={"query": PRINTAVO_ORDERS_QUERY,
                  "variables": {"after": after, "statusIds": status_ids}},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if "errors" in data:
            raise RuntimeError(f"Printavo GraphQL error: {data['errors']}")
        conn = data["data"]["orders"]
        nodes.extend(conn["nodes"])
        if conn["pageInfo"]["hasNextPage"]:
            after = conn["pageInfo"]["endCursor"]
        else:
            break
    return nodes


# ----------------------------------------------------------------------------
# 2. Aggregate demand + per-order records
# ----------------------------------------------------------------------------

@dataclass
class BlankDemand:
    # key = (style_number, color) ; value = {size_label: qty}
    rows: dict = field(default_factory=lambda: defaultdict(lambda: defaultdict(int)))
    descriptions: dict = field(default_factory=dict)

    def add(self, style, color, size_label, qty, desc=""):
        if qty:
            self.rows[(style, color)][size_label] += int(qty)
            if desc:
                self.descriptions[(style, color)] = desc


def _iter_line_items(order):
    """Yield line item dicts regardless of minor schema nesting differences."""
    for g in (order.get("lineItemGroups", {}) or {}).get("nodes", []) or []:
        for li in (g.get("lineItems", {}) or {}).get("nodes", []) or []:
            yield li


def _normalize_sizes(li):
    """Return [(size_label, qty)] from either v2 `sizes[]` or legacy size_* keys."""
    out = []
    if isinstance(li.get("sizes"), list):                # v2 shape
        for s in li["sizes"]:
            label = SIZE_ALIASES.get(str(s.get("size", "")).lower().strip())
            if label:
                out.append((label, s.get("quantity", 0)))
    else:                                                # legacy flat shape
        for k, v in li.items():
            label = SIZE_ALIASES.get(k.lower())
            if label and v:
                out.append((label, v))
    return out


def aggregate_demand(orders) -> BlankDemand:
    demand = BlankDemand()
    for o in orders:
        for li in _iter_line_items(o):
            style = (li.get("styleNumber") or li.get("style_number") or "").strip()
            color = (li.get("color") or "").strip()
            desc = (li.get("styleDescription") or li.get("style_description") or "").strip()
            if not style:
                continue
            for size_label, qty in _normalize_sizes(li):
                demand.add(style, color, size_label, qty, desc)
    return demand


def extract_order_records(orders, blanks_task_name):
    """Per-order info for write-back: id, visualId, the (style,color) keys it
    contains, and the ID of its open 'Order Blanks' task (if any)."""
    records = []
    want = blanks_task_name.strip().lower()
    for o in orders:
        keys = set()
        for li in _iter_line_items(o):
            style = (li.get("styleNumber") or li.get("style_number") or "").strip()
            color = (li.get("color") or "").strip()
            if style:
                keys.add((style, color))
        task_id = None
        for t in (o.get("tasks", {}) or {}).get("nodes", []) or []:
            if (t.get("name") or "").strip().lower() == want and not t.get("completed"):
                task_id = t.get("id")
                break
        records.append({"id": o.get("id"), "visualId": o.get("visualId"),
                        "keys": keys, "blanks_task_id": task_id})
    return records


# ----------------------------------------------------------------------------
# 3. SKU map: Printavo (style, color) -> SanMar (style, color code)
# ----------------------------------------------------------------------------

def load_sku_map(path="sku_map.csv"):
    """CSV columns: printavo_style, printavo_color, sanmar_style, sanmar_color, confirmed
    `confirmed` = yes/no. Unconfirmed rows are flagged in the report, never ordered."""
    m = {}
    if not Path(path).exists():
        return m
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            key = (r["printavo_style"].strip(), r["printavo_color"].strip())
            m[key] = {
                "sanmar_style": r["sanmar_style"].strip(),
                "sanmar_color": r["sanmar_color"].strip(),
                "confirmed": r.get("confirmed", "no").strip().lower() == "yes",
            }
    return m


def map_to_sanmar(demand: BlankDemand, sku_map):
    """Returns (po_lines, unmapped). po_lines are confirmed-mappable SKUs only."""
    po_lines, unmapped = [], []
    for (style, color), sizes in demand.rows.items():
        mp = sku_map.get((style, color))
        total = sum(sizes.values())
        base = {"printavo_style": style, "printavo_color": color,
                "desc": demand.descriptions.get((style, color), ""),
                "sizes": dict(sizes), "total": total}
        if not mp:
            unmapped.append({**base, "reason": "No SKU map entry"})
        elif not mp["confirmed"]:
            unmapped.append({**base, "reason": "Mapping not confirmed",
                             "sanmar_style": mp["sanmar_style"],
                             "sanmar_color": mp["sanmar_color"]})
        else:
            po_lines.append({**base, "sanmar_style": mp["sanmar_style"],
                             "sanmar_color": mp["sanmar_color"]})
    return po_lines, unmapped


# ----------------------------------------------------------------------------
# 4. Report (XLSX)
# ----------------------------------------------------------------------------

def build_report(demand, po_lines, unmapped, day, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(bold=True, color="FFFFFF")
    warn_fill = PatternFill("solid", fgColor="FDE68A")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = hdr_fill, hdr_font
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: blanks to order (confirmed)
    ws = wb.active
    ws.title = "Blanks To Order"
    cols = ["SanMar Style", "SanMar Color", "Description"] + SIZE_ORDER + ["TOTAL"]
    ws.append(cols)
    style_header(ws, len(cols))
    grand = 0
    for ln in sorted(po_lines, key=lambda x: (x["sanmar_style"], x["sanmar_color"])):
        row = [ln["sanmar_style"], ln["sanmar_color"], ln["desc"]]
        row += [ln["sizes"].get(s, "") for s in SIZE_ORDER]
        row += [ln["total"]]
        ws.append(row)
        grand += ln["total"]
    ws.append([])
    ws.append(["", "", "GRAND TOTAL UNITS"] + [""] * len(SIZE_ORDER) + [grand])
    for col in "ABC":
        ws.column_dimensions[col].width = 22

    # Sheet 2: needs attention (unmapped / unconfirmed) -- do NOT order these blind
    ws2 = wb.create_sheet("Needs Attention")
    cols2 = ["Printavo Style", "Printavo Color", "Description",
             "Suggested SanMar Style", "Suggested SanMar Color", "Reason", "Total Units"]
    ws2.append(cols2)
    style_header(ws2, len(cols2))
    for u in unmapped:
        ws2.append([u["printavo_style"], u["printavo_color"], u["desc"],
                    u.get("sanmar_style", ""), u.get("sanmar_color", ""),
                    u["reason"], u["total"]])
        for c in range(1, len(cols2) + 1):
            ws2.cell(row=ws2.max_row, column=c).fill = warn_fill
    for col in "ABCDE":
        ws2.column_dimensions[col].width = 22
    ws2.column_dimensions["F"].width = 26

    wb.save(out_path)
    return grand


# ----------------------------------------------------------------------------
# 5. Email
# ----------------------------------------------------------------------------

def email_report(cfg, out_path, day, summary):
    e = cfg.get("email")
    if not e or not e.get("enabled", False):
        print("[email] disabled in config -- skipping send")
        return
    msg = EmailMessage()
    msg["Subject"] = f"Blanks To Order -- {day.isoformat()}"
    msg["From"] = e["from"]
    msg["To"] = ", ".join(e["to"]) if isinstance(e["to"], list) else e["to"]
    msg.set_content(summary)
    with open(out_path, "rb") as f:
        msg.add_attachment(f.read(), maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename=Path(out_path).name)
    with smtplib.SMTP(e["smtp_host"], e["smtp_port"]) as s:
        s.starttls()
        s.login(e["smtp_user"], e["smtp_pass"])
        s.send_message(msg)
    print(f"[email] sent to {msg['To']}")


# ----------------------------------------------------------------------------
# 6. Printavo write-back (status move + task complete)
# ----------------------------------------------------------------------------

def printavo_mutate(cfg, query, variables):
    resp = requests.post(PRINTAVO_ENDPOINT, headers=printavo_headers(cfg),
                         json={"query": query, "variables": variables}, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if "errors" in data:
        raise RuntimeError(f"Printavo mutation error: {data['errors']}")
    return data["data"]


STATUS_UPDATE_MUT = """
mutation($parentId: ID!, $statusId: ID!) {
  statusUpdate(parentId: $parentId, statusId: $statusId) {
    ... on Invoice { id visualId }
    ... on Quote { id visualId }
  }
}"""

TASK_COMPLETE_MUT = """
mutation($id: ID!) {
  taskUpdate(id: $id, input: { completed: true }) { id name completed }
}"""


def advance_orders(cfg, records, po_lines, dry_run=True):
    """Move fully-ordered orders to AWAITING GARMENTS and check 'Order Blanks'.

    SAFETY: only advances an order if EVERY (style,color) on it is in the
    confirmed po_lines. An order with any unmapped/held line is left in place,
    so it stays in the pull bucket until its blanks are actually ordered.
    """
    confirmed_keys = {(l["printavo_style"], l["printavo_color"]) for l in po_lines}
    awaiting = str(cfg["printavo"].get("awaiting_garments_status_id", "")).strip()
    if not awaiting:
        print("[writeback] awaiting_garments_status_id not set -- skipping status moves")

    advanced, held = [], []
    for r in records:
        fully_covered = r["keys"] and r["keys"].issubset(confirmed_keys)
        if not fully_covered:
            held.append(r["visualId"])
            continue
        if dry_run:
            print(f"[writeback] DRY RUN -- would advance #{r['visualId']} -> AWAITING "
                  f"GARMENTS + complete Order Blanks task ({r['blanks_task_id']})")
            advanced.append(r["visualId"])
            continue
        if awaiting:
            printavo_mutate(cfg, STATUS_UPDATE_MUT,
                            {"parentId": r["id"], "statusId": awaiting})
        if r["blanks_task_id"]:
            printavo_mutate(cfg, TASK_COMPLETE_MUT, {"id": r["blanks_task_id"]})
        print(f"[writeback] #{r['visualId']} -> AWAITING GARMENTS, Order Blanks checked")
        advanced.append(r["visualId"])
    if held:
        print(f"[writeback] held (unordered lines remain): {', '.join(map(str, held))}")
    return advanced, held


# ----------------------------------------------------------------------------
# 7. SanMar PO (gated, dry-run by default)
# ----------------------------------------------------------------------------

def sanmar_check_stock(cfg, po_lines):
    """getPreSubmitInfo per SanMar recommendation -- annotate lines w/ availability.

    TODO(next-builder)[B]: wire against the live WSDL and add an "Available"
    column to build_report(). Use zeep type factories from the WSDL; do not
    hand-build the SOAP body. Left raising so it can't be silently wrong.
    """
    from zeep import Client
    Client(cfg["sanmar"]["po_wsdl"])  # placeholder so the import/connection is real
    raise NotImplementedError(
        "TODO[B]: implement getPreSubmitInfo against the live PromoStandards WSDL.")


def submit_sanmar_po(cfg, po_lines, day, dry_run=True):
    """Builds the SanMar PromoStandards PO. DRY RUN prints the payload, sends nothing.
    Live submit requires --submit AND --i-approve AND sanmar.po_enabled: true."""
    sm = cfg["sanmar"]
    po_number = f"PCA-{day.strftime('%Y%m%d')}"
    lines = []
    for i, ln in enumerate(po_lines, start=1):
        for size in SIZE_ORDER:
            q = ln["sizes"].get(size)
            if q:
                lines.append({"lineNumber": i, "style": ln["sanmar_style"],
                              "color": ln["sanmar_color"], "size": size, "qty": q})

    # NOTE: this dict is a HUMAN-READABLE placeholder for the DRY RUN preview only.
    # TODO(next-builder)[A]: the LIVE path below must build the real request from
    # the WSDL via zeep type factories, not from this dict.
    payload = {
        "wsVersion": "2.0.0",
        "id": sm.get("username"),
        "password": "***",
        "PO": {
            "poNumber": po_number,
            "orderType": "Blank",
            "shipment": {"shipToName": sm["ship_to_name"], "address": sm["ship_to_address"]},
            "lineItems": lines,
        },
    }
    if dry_run or not sm.get("po_enabled", False):
        print(f"[sanmar] DRY RUN -- PO {po_number}, {len(lines)} line(s). Nothing sent.")
        for l in lines:
            print(f"   {l['style']:>10}  {l['color']:<18} {l['size']:<4} x {l['qty']}")
        return {"dry_run": True, "po_number": po_number, "lines": lines, "payload": payload}

    # --- LIVE PATH ----------------------------------------------------------
    # TODO(next-builder)[A]: replace the call below with the real PromoStandards
    # sendPO request built from the WSDL. Confirm operation name + arg shape on
    # the TEST WSDL first (test-ws.sanmar.com). Example skeleton:
    #
    #   from zeep import Client
    #   from zeep.wsse.username import UsernameToken
    #   client = Client(sm["po_wsdl"], wsse=UsernameToken(sm["username"], sm["password"]))
    #   PO = client.get_type("ns0:PO")(...)            # build typed PO object
    #   result = client.service.sendPO(wsVersion="2.0.0", id=sm["username"],
    #                                   password=sm["password"], PO=PO)
    #
    raise NotImplementedError(
        "TODO[A]: implement the live SanMar sendPO against the WSDL. "
        f"Prepared PO {po_number} with {len(lines)} line(s).")


# ----------------------------------------------------------------------------
# Orchestrator
# ----------------------------------------------------------------------------

def run_once(cfg, day, check_stock=False, submit=False, approved=False):
    status_ids = cfg["printavo"].get("status_ids", [])
    print(f"== Blanks bot -- snapshot {day.isoformat()} -- status_ids={status_ids} ==")
    orders = fetch_printavo_orders(cfg, status_ids)
    print(f"[printavo] {len(orders)} order(s) in target status")
    demand = aggregate_demand(orders)
    records = extract_order_records(
        orders, cfg["printavo"].get("order_blanks_task_name", "Order Blanks"))
    sku_map = load_sku_map(cfg.get("sku_map_path", "sku_map.csv"))
    po_lines, unmapped = map_to_sanmar(demand, sku_map)

    out_dir = Path(cfg.get("output_dir", "."))
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"blanks_to_order_{day.strftime('%Y%m%d')}.xlsx"
    grand = build_report(demand, po_lines, unmapped, day, out_path)

    summary = (f"{day.isoformat()} blanks run\n"
               f"Confirmed SKUs to order: {len(po_lines)}  ({grand} units)\n"
               f"Needs attention (unmapped/unconfirmed): {len(unmapped)}\n"
               f"Sheet attached.")
    print(summary)
    email_report(cfg, out_path, day, summary)

    if check_stock and po_lines:
        sanmar_check_stock(cfg, po_lines)

    if submit:
        if not approved:
            print("[sanmar] --submit given without --i-approve. Refusing to buy. "
                  "DRY RUN only.")
        live = submit and approved and cfg["sanmar"].get("po_enabled", False)
        submit_sanmar_po(cfg, po_lines, day, dry_run=not live)
        # Advance Printavo only when the PO actually went live; otherwise preview.
        advance_orders(cfg, records, po_lines, dry_run=not live)

    return out_path


def main():
    p = argparse.ArgumentParser(description="Printavo -> SanMar blanks ordering bot")
    p.add_argument("--init", action="store_true",
                   help="write config.example.yaml + sku_map.csv, then exit")
    p.add_argument("--once", action="store_true", help="run a single accumulation cycle")
    p.add_argument("--list-statuses", action="store_true",
                   help="print Printavo statuses + IDs, then exit")
    p.add_argument("--date", help="snapshot label YYYY-MM-DD (default: today)")
    p.add_argument("--config", default="config.yaml")
    p.add_argument("--check-stock", action="store_true",
                   help="(after TODO[B]) add SanMar inventory column")
    p.add_argument("--submit", action="store_true", help="attempt SanMar PO submit")
    p.add_argument("--i-approve", action="store_true",
                   help="explicit human approval required to spend money")
    args = p.parse_args()

    if args.init:
        write_init_files()
        return

    cfg = load_config(args.config)
    day = dt.date.fromisoformat(args.date) if args.date else dt.date.today()

    if args.list_statuses:
        list_statuses(cfg)
    elif args.once:
        run_once(cfg, day, check_stock=args.check_stock,
                 submit=args.submit, approved=args.i_approve)
    else:
        p.print_help()


if __name__ == "__main__":
    main()
